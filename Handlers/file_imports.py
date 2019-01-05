import openpyxl


"""Master Upload File"""
Registration_file = '/Users/nbenzschawel/Downloads/SEVIS_Reg/2019/Spring/SEVIS Registrion Spring 2019.xlsx'


""" COL Student Data """
COL_students = '/Users/nbenzschawel/Downloads/SEVIS_Reg/2019/Spring/Raw_Files/Change_of_Level_(COL)_Students.xlsx'


""" New Student Data """
new_stud_issm_data = '/Users/nbenzschawel/Downloads/SEVIS_Reg/2019/Spring/Raw_Files/SEVIS_Registration_Tracking-New_Students-SEVIS_Pending.xlsx'
sevis_inital_student_data = '/Users/nbenzschawel/Downloads/SEVIS_Reg/2019/Spring/Raw_Files/SEVIS_raw_data.xlsx'

wb1_new = openpyxl.load_workbook(new_stud_issm_data) # workbook from ISSM
wb2_new = openpyxl.load_workbook(sevis_inital_student_data) # workbook from SEVIS RTI

ws1_new = wb1_new.active
ws2_new = wb2_new.active


""" SEVIS Active Student Data """
active_stud_issm_data = '/Users/nbenzschawel/Downloads/SEVIS_Reg/2019/Spring/Raw_Files/All_SEVIS-Active_Student_Tracking.xlsx'
active_student_req_reg = '/Users/nbenzschawel/Downloads/SEVIS_Reg/2019/Spring/Raw_Files/Fall 201840 - Registration.xlsx'

wb2_active = openpyxl.load_workbook(active_student_req_reg)
sheet = wb2_active.worksheets[4]
ws2 = wb2_active.active

wb1 = openpyxl.load_workbook(active_stud_issm_data)
ws = wb1.active


""" Transfer Student Data """
new_transfer_data = '/Users/nbenzschawel/Downloads/SEVIS_Reg/2019/Spring/Raw_Files/New_sevis_data.xlsx'
current_transfer_data = '/Users/nbenzschawel/Downloads/SEVIS_Reg/2019/Spring/Raw_Files/SEVIS_transfer_data.xlsx'

wb_trans = openpyxl.load_workbook(new_transfer_data)
wb2_trans = openpyxl.load_workbook(current_transfer_data)
old_sheet = wb_trans.worksheets[0]
new_sheet = wb2_trans.worksheets[1]
final_sheet = wb2_trans.worksheets[0]

row_max_old = old_sheet.max_row
col_max_old = old_sheet.max_column

row_max_current = new_sheet.max_row
col_max_current = new_sheet.max_column

row_max_final = final_sheet.max_row
col_max_final = final_sheet.max_column


""" No Show-Cancels for Admissions"""
SEVIS_initial_status_stud = '/Users/nbenzschawel/Downloads/SEVIS_Reg/2019/Spring/Raw_Files/SEVIS - Initial Status Students.xlsx'
No_show_UG = '/Users/nbenzschawel/Downloads/SEVIS_Reg/2019/Spring/Raw_Files/No-Shows-UG.xlsx'
No_show_GR = '/Users/nbenzschawel/Downloads/SEVIS_Reg/2019/Spring/Raw_Files/No-Shows-Grad.xlsx'

SEV_initial_data = openpyxl.load_workbook(SEVIS_initial_status_stud)
wb_cancel_ug = openpyxl.load_workbook(No_show_UG)
wb_cancel_gr = openpyxl.load_workbook(No_show_GR)

ug_sheet = wb_cancel_ug.worksheets[0]
gr_sheet = wb_cancel_gr.worksheets[0]

ug_row_max = ug_sheet.max_row
ug_col_max = ug_sheet.max_column
gr_row_max = gr_sheet.max_row
gr_col_max = gr_sheet.max_column

ws_cancel_ug = wb_cancel_ug.active
ws_cancel_gr = wb_cancel_gr.active


""" Completed Student Data """
comp_student_cleanup_report = '/Users/nbenzschawel/Downloads/SEVIS_Reg/2019/Spring/Raw_Files/Completed_Students_Cleanup_Report_201840.xlsx'
sevis_completed_students = '/Users/nbenzschawel/Downloads/SEVIS_Reg/2019/Spring/Raw_Files/SEVIS - Completed Status Students (past 18 months).xlsx'

completed_student_wb2 = openpyxl.load_workbook(comp_student_cleanup_report)

comp_sheet = completed_student_wb2.worksheets[0]

comp_ws = completed_student_wb2.active

wb1_complete = openpyxl.load_workbook(sevis_completed_students)

ws1 = wb1_complete.active


""" Graduated Student Data """
graduated_students = '/Users/nbenzschawel/Downloads/SEVIS_Reg/2019/Spring/Raw_Files/201820_201830_graduated_students.xlsx'
wb1_grad = openpyxl.load_workbook(active_stud_issm_data)

sheet_grad = wb1_grad.worksheets[0]
ws_grad = wb1_grad.active

# Think about adding a try, except function here to check the file directory.
wb2_grad = openpyxl.load_workbook('/Users/nbenzschawel/Downloads/SEVIS_Reg/2019/Spring/Raw_Files/201840_graduated_students_test.xlsx')
wb3Live_grad = openpyxl.load_workbook(graduated_students)

sheet2_grad = wb2_grad.worksheets[0]
sheet3_grad = wb3Live_grad.worksheets[0]

ws2_grad = wb2_grad.active
ws3_grad = wb3Live_grad.active


"""SEVIS Registration Timelines"""
current_registration_timeline = '/Volumes/ISS/Staff Only/Student Lists/SEVIS Registration/SEVIS Registration Timelines/SEVIS REGISTRATION 19S.docx'